import os
import re
import json
import sys
from pathlib import Path
from datetime import datetime, timezone

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
ENV_PATH = ROOT / ".env.supabase.local"
WORKBOOK_PATH = ROOT / "HRI Dashbard Map.xlsx"
DRY_RUN = "--dry-run" in sys.argv

load_dotenv(ENV_PATH)

SUPABASE_URL = os.getenv("SUPABASE_URL", "").strip().rstrip("/")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()

if not DRY_RUN and (not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY):
    raise SystemExit("Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY in .env.supabase.local")

if not WORKBOOK_PATH.exists():
    raise SystemExit(f"Workbook not found: {WORKBOOK_PATH}")

REST_URL = f"{SUPABASE_URL}/rest/v1" if SUPABASE_URL else ""

DEPARTMENTS = [
    {"department_id": "executive_home", "parent_department_id": None, "department_name": "Executive Home", "level": "home", "sort_order": 1},
    {"department_id": "employee_home", "parent_department_id": None, "department_name": "Employee Home", "level": "home", "sort_order": 2},
    {"department_id": "program_management", "parent_department_id": None, "department_name": "Program Management", "level": "department", "sort_order": 10},
    {"department_id": "heavy_industrial", "parent_department_id": "program_management", "department_name": "Heavy Industrial", "level": "subdepartment", "sort_order": 11},
    {"department_id": "cold_storage_distribution", "parent_department_id": "program_management", "department_name": "Cold Storage & Distribution", "level": "subdepartment", "sort_order": 12},
    {"department_id": "food_beverage", "parent_department_id": "program_management", "department_name": "Food & Beverage", "level": "subdepartment", "sort_order": 13},
    {"department_id": "marketing", "parent_department_id": "program_management", "department_name": "Marketing", "level": "subdepartment", "sort_order": 14},
    {"department_id": "cost_analytics", "parent_department_id": "program_management", "department_name": "Cost Analytics", "level": "subdepartment", "sort_order": 15},
    {"department_id": "gc_ops", "parent_department_id": None, "department_name": "GC Operations", "level": "department", "sort_order": 20},
    {"department_id": "gc_ops_pm", "parent_department_id": "gc_ops", "department_name": "GC Ops - PM", "level": "subdepartment", "sort_order": 21},
    {"department_id": "gc_ops_pe", "parent_department_id": "gc_ops", "department_name": "GC Ops - PE", "level": "subdepartment", "sort_order": 22},
    {"department_id": "gc_ops_super", "parent_department_id": "gc_ops", "department_name": "GC Ops - Superintendent", "level": "subdepartment", "sort_order": 23},
    {"department_id": "steel_thermal", "parent_department_id": None, "department_name": "Steel & Thermal", "level": "department", "sort_order": 30},
    {"department_id": "preconstruction", "parent_department_id": None, "department_name": "Preconstruction", "level": "department", "sort_order": 40},
    {"department_id": "design_engineering", "parent_department_id": "preconstruction", "department_name": "Design / Engineering", "level": "subdepartment", "sort_order": 41},
    {"department_id": "estimating", "parent_department_id": "preconstruction", "department_name": "Estimating", "level": "subdepartment", "sort_order": 42},
    {"department_id": "compliance", "parent_department_id": None, "department_name": "EHS-QA / Compliance", "level": "department", "sort_order": 50},
    {"department_id": "ehs", "parent_department_id": "compliance", "department_name": "EHS", "level": "subdepartment", "sort_order": 51},
    {"department_id": "project_controls", "parent_department_id": "compliance", "department_name": "Project Controls", "level": "subdepartment", "sort_order": 52},
    {"department_id": "administration", "parent_department_id": None, "department_name": "Administration", "level": "department", "sort_order": 60},
    {"department_id": "finance", "parent_department_id": "administration", "department_name": "Finance", "level": "subdepartment", "sort_order": 61},
    {"department_id": "human_resources", "parent_department_id": "administration", "department_name": "Human Resources", "level": "subdepartment", "sort_order": 62},
    {"department_id": "accounting", "parent_department_id": "administration", "department_name": "Accounting", "level": "subdepartment", "sort_order": 63},
    {"department_id": "it", "parent_department_id": None, "department_name": "Information Technology", "level": "department", "sort_order": 70},
    {"department_id": "project_pages", "parent_department_id": None, "department_name": "Project Pages", "level": "project_area", "sort_order": 80},
]

RAW_DEPARTMENT_MAP = {
    "program mgt": ("program_management", None),
    "progam mgt": ("program_management", None),
    "program management": ("program_management", None),
    "program mgt-marketing": ("program_management", "marketing"),
    "program mgt-cost analytics": ("program_management", "cost_analytics"),
    "program mgt-cs & d": ("program_management", "cold_storage_distribution"),
    "program mgt- cs & d": ("program_management", "cold_storage_distribution"),
    "progam mgt-cs & d": ("program_management", "cold_storage_distribution"),
    "progam mgt-f&b": ("program_management", "food_beverage"),
    "progam mgt-f & b": ("program_management", "food_beverage"),
    "progam mgt-hi": ("program_management", "heavy_industrial"),
    "gc ops": ("gc_ops", None),
    "gc ops-pm": ("gc_ops", "gc_ops_pm"),
    "gc ops-pe": ("gc_ops", "gc_ops_pe"),
    "gc ops- super": ("gc_ops", "gc_ops_super"),
    "s & t": ("steel_thermal", None),
    "steel & thermal": ("steel_thermal", None),
    "preconstruction": ("preconstruction", None),
    "preconstruction-design": ("preconstruction", "design_engineering"),
    "preconstruction-estimating": ("preconstruction", "estimating"),
    "compliance": ("compliance", None),
    "compliance-ehs": ("compliance", "ehs"),
    "compliance-pcc": ("compliance", "project_controls"),
    "f & a": ("administration", None),
    "f & a-accounting": ("administration", "accounting"),
    "f & a-hr": ("administration", "human_resources"),
    "it": ("it", None),
}

PAGE_TO_DEPT = {
    "program mgt": "program_management",
    "gc ops": "gc_ops",
    "steel & thermal": "steel_thermal",
    "s & t": "steel_thermal",
    "preconstruction": "preconstruction",
    "compliance": "compliance",
    "administration": "administration",
    "it": "it",
}

ACCESS_LEVEL_MAP = {
    "executive": {"roles": ["executive"], "departments": ["all"], "subdepartments": ["all"], "landing_page": "executive_home"},
    "program mgt": {"roles": ["program_management"], "departments": ["program_management"], "subdepartments": ["all_program_management"], "landing_page": "department_home"},
    "gc ops": {"roles": ["gc_ops"], "departments": ["gc_ops"], "subdepartments": [], "landing_page": "department_home"},
    "s & t": {"roles": ["steel_thermal"], "departments": ["steel_thermal"], "subdepartments": [], "landing_page": "department_home"},
    "design": {"roles": ["design"], "departments": ["preconstruction"], "subdepartments": ["design_engineering", "estimating"], "landing_page": "department_home"},
    "f & b": {"roles": ["food_beverage"], "departments": ["program_management"], "subdepartments": ["food_beverage"], "landing_page": "department_home"},
    "hi": {"roles": ["heavy_industrial"], "departments": ["program_management"], "subdepartments": ["heavy_industrial"], "landing_page": "department_home"},
    "cs & d": {"roles": ["cold_storage_distribution"], "departments": ["program_management"], "subdepartments": ["cold_storage_distribution"], "landing_page": "department_home"},
    "marketing": {"roles": ["marketing"], "departments": ["program_management"], "subdepartments": ["marketing"], "landing_page": "department_home"},
    "cost analytics": {"roles": ["cost_analytics"], "departments": ["program_management"], "subdepartments": ["cost_analytics"], "landing_page": "department_home"},
    "finance": {"roles": ["finance"], "departments": ["administration"], "subdepartments": ["finance", "accounting"], "landing_page": "department_home"},
    "hr": {"roles": ["human_resources"], "departments": ["administration"], "subdepartments": ["human_resources"], "landing_page": "department_home"},
    "ehs": {"roles": ["ehs"], "departments": ["compliance"], "subdepartments": ["ehs"], "landing_page": "department_home"},
    "pcc": {"roles": ["project_controls"], "departments": ["compliance"], "subdepartments": ["project_controls"], "landing_page": "department_home"},
    "it": {"roles": ["it"], "departments": ["it"], "subdepartments": [], "landing_page": "department_home"},
}

CARD_ACCESS_MAP = {
    "individual": ("individual", "self"),
    "all": ("all", "all"),
    "all ehs": ("group", "ehs"),
    "f & b": ("group", "food_beverage"),
    "hi": ("group", "heavy_industrial"),
    "cs & d": ("group", "cold_storage_distribution"),
    "s & t": ("group", "steel_thermal"),
    "finance": ("group", "finance"),
    "pcc": ("group", "project_controls"),
}

PROJECT_PROGRAM_MAP = {
    "cs & d": "cold_storage_distribution",
    "f & b": "food_beverage",
    "hi": "heavy_industrial",
    "s & t": "steel_thermal",
}

ROLE_TOKEN_MAP = {
    "executive": ("role", "executive"),
    "program": ("program", "project_program"),
    "program & executive": ("compound", "program_executive"),
    "gc ops": ("role", "gc_ops"),
    "s & t": ("role", "steel_thermal"),
    "design": ("role", "design"),
    "f & b": ("program", "food_beverage"),
    "hi": ("program", "heavy_industrial"),
    "cs & d": ("program", "cold_storage_distribution"),
    "finance": ("role", "finance"),
    "hr": ("role", "human_resources"),
    "ehs": ("role", "ehs"),
    "pcc": ("role", "project_controls"),
    "it": ("role", "it"),
}

IGNORED_VISIBLE_EMPLOYEE_TEXT = {
    "",
    "list specific outdide of executives and program individuals",
    "list specific outside of executives and program individuals",
    "none",
    "n/a",
}

def clean(value):
    if value is None:
        return ""
    return str(value).strip()

def clean_lower(value):
    return re.sub(r"\s+", " ", clean(value).lower())

def slug(value):
    value = clean_lower(value)
    value = value.replace("&", "and")
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_") or "unknown"

def as_int(value):
    if value is None or clean(value) == "":
        return None
    try:
        return int(float(value))
    except Exception:
        return None

def is_yes(value):
    return clean_lower(value) in {"yes", "y", "true", "1"}

def is_active(value):
    return clean_lower(value) == "active"

def split_tokens(value):
    return [t.strip() for t in re.split(r"[,;\n]+", clean(value)) if t.strip()]

def normalize_email(value):
    return clean_lower(value)

def map_raw_department(raw):
    return RAW_DEPARTMENT_MAP.get(clean_lower(raw), (None, None))

def map_access_level(access_level):
    return ACCESS_LEVEL_MAP.get(clean_lower(access_level), {"roles": [], "departments": [], "subdepartments": [], "landing_page": "employee_home"})

def map_card_access(card_access_raw):
    raw = clean_lower(card_access_raw)
    if raw in CARD_ACCESS_MAP:
        return CARD_ACCESS_MAP[raw]
    if raw:
        return ("group", slug(raw))
    return ("none", "none")

def sheet_rows(ws):
    headers = []
    blank_count = 0
    for idx, cell in enumerate(ws[1], start=1):
        header = clean(cell.value)
        if not header:
            blank_count += 1
            header = f"__blank_{idx}"
        headers.append(header)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(clean(v) for v in row):
            continue
        yield {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}

def supabase_request(method, table, params="", payload=None, prefer=None):
    if DRY_RUN:
        return None
    url = f"{REST_URL}/{table}{params}"
    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    if prefer:
        headers["Prefer"] = prefer
    response = requests.request(method, url, headers=headers, json=payload, timeout=60)
    if response.status_code >= 400:
        raise RuntimeError(f"Supabase {method} {table} failed: {response.status_code} {response.text[:1000]}")
    if response.text:
        try:
            return response.json()
        except Exception:
            return response.text
    return None

def reset_rules():
    for table, column in [
        ("hri_employee_dashboard_access", "employee_email"),
        ("hri_employee_card_access", "employee_email"),
        ("hri_project_access_rules", "project_id"),
    ]:
        supabase_request("DELETE", table, params=f"?{column}=neq.__never__", prefer="return=minimal")

def upsert_rows(table, rows, on_conflict):
    if not rows:
        return
    for i in range(0, len(rows), 500):
        chunk = rows[i:i+500]
        supabase_request(
            "POST",
            table,
            params=f"?on_conflict={on_conflict}",
            payload=chunk,
            prefer="resolution=merge-duplicates,return=minimal",
        )

def insert_rows(table, rows):
    if not rows:
        return
    for i in range(0, len(rows), 500):
        supabase_request("POST", table, payload=rows[i:i+500], prefer="return=minimal")

def build_artifacts(ws):
    rows = []
    for row in sheet_rows(ws):
        title = clean(row.get("Tile title"))
        page = clean(row.get("Page"))
        if not title or not page:
            continue
        department_id = PAGE_TO_DEPT.get(clean_lower(page), slug(page))
        tile_number = as_int(row.get("Tile number on page"))
        artifact_url = clean(row.get("Artifact Link"))
        artifact_id = slug(f"{department_id}_{title}")
        artifact_type = "powerbi" if "powerbi.com" in artifact_url.lower() else "claude_artifact" if "claude.ai" in artifact_url.lower() else "link"
        rows.append({
            "artifact_id": artifact_id,
            "tile_title": title,
            "tile_number": tile_number,
            "department_id": department_id,
            "storage_key": clean(row.get("Storage Key")) or None,
            "stats_shown": clean(row.get("Stats Shown (not in use)")) or None,
            "source_name": clean(row.get("Source HTML/Project")) or None,
            "artifact_url": artifact_url or None,
            "artifact_type": artifact_type,
            "status": "active" if artifact_url else "coming_soon",
            "sort_order": tile_number or 999,
            "raw_page": page,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        })
    return rows

def build_employees_and_access(ws):
    employees_by_email = {}
    dashboard_access = set()
    card_access = set()
    duplicate_counts = {}

    for row in sheet_rows(ws):
        email = normalize_email(row.get("Email address"))
        if not email:
            continue

        duplicate_counts[email] = duplicate_counts.get(email, 0) + 1
        name = clean(row.get("Employee_Name"))
        access_level = clean(row.get("Access Level"))
        active_text = clean(row.get("Active/Inactive"))
        invite_text = clean(row.get("Current Invite"))
        card_raw = clean(row.get("Accesable Employee Cards"))
        department_id, subdepartment_id = map_raw_department(row.get("Department"))
        access = map_access_level(access_level)
        can_login = is_active(active_text) and is_yes(invite_text) and clean_lower(access_level) != "none"

        existing = employees_by_email.get(email)
        if existing:
            departments = set(existing.get("_departments", []))
            subdepartments = set(existing.get("_subdepartments", []))
            roles = set(existing.get("_roles", []))
            departments.update([d for d in access["departments"] if d])
            subdepartments.update([s for s in access["subdepartments"] if s])
            roles.update([r for r in access["roles"] if r])
            if department_id:
                departments.add(department_id)
            if subdepartment_id:
                subdepartments.add(subdepartment_id)
            existing["_departments"] = sorted(departments)
            existing["_subdepartments"] = sorted(subdepartments)
            existing["_roles"] = sorted(roles)
            existing["can_login"] = existing["can_login"] or can_login
            existing["is_current_invite"] = existing["is_current_invite"] or is_yes(invite_text)
            existing["is_active"] = existing["is_active"] or is_active(active_text)
            existing["duplicate_note"] = "Merged duplicate workbook rows"
        else:
            employees_by_email[email] = {
                "employee_email": email,
                "employee_name": name,
                "title": clean(row.get("Title")) or None,
                "employee_code": clean(row.get("Employee_Code")) or None,
                "department_raw": clean(row.get("Department")) or None,
                "department_id": department_id,
                "subdepartment_id": subdepartment_id,
                "active_status": active_text or None,
                "current_invite": invite_text or None,
                "access_level": access_level or None,
                "employee_card_access_raw": card_raw or None,
                "is_active": is_active(active_text),
                "is_current_invite": is_yes(invite_text),
                "can_login": can_login,
                "landing_page": access["landing_page"],
                "duplicate_note": None,
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "_departments": sorted(set([d for d in access["departments"] if d] + ([department_id] if department_id else []))),
                "_subdepartments": sorted(set([s for s in access["subdepartments"] if s] + ([subdepartment_id] if subdepartment_id else []))),
                "_roles": sorted(set([r for r in access["roles"] if r])),
            }

        employee = employees_by_email[email]
        for role in employee.get("_roles", []):
            dashboard_access.add((email, "role", role))
        for department in employee.get("_departments", []):
            dashboard_access.add((email, "department", department))
        for subdepartment in employee.get("_subdepartments", []):
            dashboard_access.add((email, "subdepartment", subdepartment))
        if "all" in employee.get("_departments", []) or "executive" in employee.get("_roles", []):
            dashboard_access.add((email, "artifact", "all"))
            dashboard_access.add((email, "project", "all"))

        scope, scope_value = map_card_access(card_raw)
        if clean_lower(access_level) == "executive":
            scope, scope_value = "all", "all"
        if clean_lower(access_level) == "hr" and scope == "none":
            scope, scope_value = "all", "all"
        card_access.add((email, scope, scope_value))

    employees = []
    for email, employee in employees_by_email.items():
        employee.pop("_departments", None)
        employee.pop("_subdepartments", None)
        employee.pop("_roles", None)
        if duplicate_counts.get(email, 0) > 1:
            employee["duplicate_note"] = "Merged duplicate workbook rows"
        employees.append(employee)

    dashboard_rows = [
        {"employee_email": email, "access_type": access_type, "access_value": value}
        for email, access_type, value in sorted(dashboard_access)
    ]
    card_rows = [
        {"employee_email": email, "scope": scope, "scope_value": value}
        for email, scope, value in sorted(card_access)
    ]
    return employees, dashboard_rows, card_rows

def parse_project_access(project_id, program_id, visible_to_raw, visible_to_employees_raw):
    rules = set()
    visible_to = clean_lower(visible_to_raw)

    if "program" in visible_to and program_id:
        rules.add((project_id, "program", program_id))
    if "executive" in visible_to:
        rules.add((project_id, "role", "executive"))

    for token in split_tokens(visible_to_raw):
        key = clean_lower(token)
        mapped = ROLE_TOKEN_MAP.get(key)
        if not mapped:
            continue
        access_type, access_value = mapped
        if access_type == "compound":
            if program_id:
                rules.add((project_id, "program", program_id))
            rules.add((project_id, "role", "executive"))
        elif access_value == "project_program" and program_id:
            rules.add((project_id, "program", program_id))
        else:
            rules.add((project_id, access_type, access_value))

    visible_employees = clean_lower(visible_to_employees_raw)
    if visible_employees not in IGNORED_VISIBLE_EMPLOYEE_TEXT:
        for token in split_tokens(visible_to_employees_raw):
            if "@" in token:
                rules.add((project_id, "employee", normalize_email(token)))
            else:
                rules.add((project_id, "employee_name_review", token))

    return [
        {"project_id": project_id, "access_type": access_type, "access_value": value}
        for project_id, access_type, value in sorted(rules)
    ]

def build_projects(ws):
    projects = []
    access_rules = []
    for row in sheet_rows(ws):
        program_code = clean(row.get("Program"))
        program_name = clean(row.get("__blank_2"))
        job_number = clean(row.get("Job Number"))
        job_name = clean(row.get("Job Name"))
        if not job_number or not job_name:
            continue
        if job_number.endswith(".0"):
            job_number = job_number[:-2]
        program_id = PROJECT_PROGRAM_MAP.get(clean_lower(program_name), slug(program_name))
        project_id = f"project_{slug(job_number)}"
        visible_to_raw = clean(row.get("Visible to"))
        visible_to_employees_raw = clean(row.get("Visible to employees"))
        projects.append({
            "project_id": project_id,
            "program_code": program_code or None,
            "program_name": program_name or None,
            "program_id": program_id or None,
            "status": clean(row.get("Status")) or None,
            "job_number": job_number,
            "job_name": job_name,
            "visible_to_raw": visible_to_raw or None,
            "visible_to_employees_raw": visible_to_employees_raw or None,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        })
        access_rules.extend(parse_project_access(project_id, program_id, visible_to_raw, visible_to_employees_raw))
    return projects, access_rules

def main():
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    required_sheets = ["Artifact Map", "EMP List", "Project Map"]
    missing = [name for name in required_sheets if name not in wb.sheetnames]
    if missing:
        raise SystemExit(f"Workbook is missing sheets: {missing}")

    artifacts = build_artifacts(wb["Artifact Map"])
    employees, dashboard_access, card_access = build_employees_and_access(wb["EMP List"])
    projects, project_access = build_projects(wb["Project Map"])

    summary = {
        "mode": "dry_run" if DRY_RUN else "supabase_sync",
        "departments": len(DEPARTMENTS),
        "artifacts": len(artifacts),
        "employees": len(employees),
        "employees_can_login_now": sum(1 for e in employees if e["can_login"]),
        "dashboard_access_rules": len(dashboard_access),
        "employee_card_access_rules": len(card_access),
        "projects": len(projects),
        "project_access_rules": len(project_access),
    }

    if DRY_RUN:
        print(json.dumps(summary, indent=2))
        print("\nDry run complete. No Supabase data was changed.")
        return

    reset_rules()
    upsert_rows("hri_departments", DEPARTMENTS, "department_id")
    upsert_rows("hri_artifacts", artifacts, "artifact_id")
    upsert_rows("hri_employees", employees, "employee_email")
    upsert_rows("hri_projects", projects, "project_id")
    insert_rows("hri_employee_dashboard_access", dashboard_access)
    insert_rows("hri_employee_card_access", card_access)
    insert_rows("hri_project_access_rules", project_access)
    insert_rows("hri_sync_log", [{
        "sync_name": "workbook_permissions_sync",
        "source_file": WORKBOOK_PATH.name,
        "summary": summary,
    }])

    print(json.dumps(summary, indent=2))
    print("\nSupabase sync complete.")

if __name__ == "__main__":
    main()
